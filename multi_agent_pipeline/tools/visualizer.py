"""
Visualizer Tools — Creates charts, Excel reports, and dashboard outputs.

Generates matplotlib charts (PNG), an Excel workbook with multiple sheets,
and returns structured visualization metadata.
"""

import csv
import json
import os
import statistics
from collections import Counter, defaultdict


def _to_float_safe(value: str) -> float | None:
    """Safely convert a string to float."""
    try:
        cleaned = str(value).strip().replace("$", "").replace(",", "").replace("€", "").replace("£", "")
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def create_visualizations(file_path: str) -> dict:
    """Create charts, Excel report, and dashboard from cleaned CSV data.

    Generates:
    1. Sales/revenue trend line chart (if date column exists)
    2. Category/region distribution pie chart
    3. Top items bar chart
    4. Multi-sheet Excel workbook (Data, Summary, Insights)
    All outputs saved to the output/ directory.

    Args:
        file_path: Path to the cleaned CSV file.

    Returns:
        A JSON string with chart file paths, Excel path, and metadata.
    """
    # Resolve path
    if not os.path.isabs(file_path):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        file_path = os.path.join(base_dir, file_path)

    if not os.path.exists(file_path):
        return json.dumps({
            "status": "error",
            "message": f"File not found: {file_path}",
            "suggestion": "Run the data cleaner first.",
        })

    # Read data
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = [row for row in reader]

    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")
    os.makedirs(output_dir, exist_ok=True)

    # Classify columns
    numeric_cols = []
    categorical_cols = []
    date_col = None

    for col in headers:
        values = [row[col].strip() for row in rows if row[col].strip()]
        if not values:
            continue
        if col.lower() in ("date", "start_date", "end_date", "order_date"):
            date_col = col
            continue
        numeric_count = sum(1 for v in values if _to_float_safe(v) is not None)
        if numeric_count / len(values) > 0.8:
            numeric_cols.append(col)
        else:
            categorical_cols.append(col)

    # Find revenue/sales column
    revenue_col = None
    for col in numeric_cols:
        if any(kw in col.lower() for kw in ("sales", "revenue", "total", "amount")):
            revenue_col = col
            break
    if not revenue_col and numeric_cols:
        revenue_col = numeric_cols[0]

    chart_files = []
    charts_metadata = []

    # ── Import matplotlib (lazy — only when tools are called) ──
    try:
        import matplotlib
        matplotlib.use("Agg")  # Non-interactive backend
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker
        has_matplotlib = True
    except ImportError:
        has_matplotlib = False

    if has_matplotlib and revenue_col:

        # Set a professional style
        plt.rcParams.update({
            "figure.facecolor": "#1a1a2e",
            "axes.facecolor": "#16213e",
            "axes.edgecolor": "#e94560",
            "axes.labelcolor": "#eee",
            "text.color": "#eee",
            "xtick.color": "#aaa",
            "ytick.color": "#aaa",
            "grid.color": "#333",
            "grid.alpha": 0.3,
            "font.size": 11,
        })

        # ── Chart 1: Trend Line Chart ──
        if date_col:
            monthly = defaultdict(float)
            monthly_count = defaultdict(int)
            for row in rows:
                date_str = row.get(date_col, "").strip()
                val = _to_float_safe(row.get(revenue_col, ""))
                if date_str and val is not None:
                    month_key = date_str[:7]
                    monthly[month_key] += val
                    monthly_count[month_key] += 1

            if monthly:
                sorted_months = sorted(monthly.keys())
                totals = [monthly[m] for m in sorted_months]

                fig, ax = plt.subplots(figsize=(12, 6))
                ax.plot(sorted_months, totals, color="#e94560", linewidth=2.5, marker="o", markersize=8, markerfacecolor="#0f3460")
                ax.fill_between(sorted_months, totals, alpha=0.15, color="#e94560")
                ax.set_title(f"Monthly {revenue_col} Trend", fontsize=16, fontweight="bold", pad=15)
                ax.set_xlabel("Month", fontsize=12)
                ax.set_ylabel(revenue_col, fontsize=12)
                ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
                ax.grid(True, linestyle="--", alpha=0.3)
                plt.xticks(rotation=45, ha="right")
                plt.tight_layout()

                chart1_path = os.path.join(output_dir, "chart_trend.png")
                plt.savefig(chart1_path, dpi=150, bbox_inches="tight")
                plt.close()
                chart_files.append(chart1_path)
                charts_metadata.append({
                    "id": "chart_1",
                    "name": f"{revenue_col} Trend",
                    "type": "line",
                    "title": f"Monthly {revenue_col} Trend",
                    "file": chart1_path,
                    "description": f"Shows {revenue_col} progression from {sorted_months[0]} to {sorted_months[-1]}",
                })

        # ── Chart 2: Distribution Pie Chart ──
        pie_col = None
        for col in categorical_cols:
            if any(kw in col.lower() for kw in ("region", "category", "type", "segment")):
                pie_col = col
                break
        if not pie_col and categorical_cols:
            pie_col = categorical_cols[0]

        if pie_col:
            cat_totals = defaultdict(float)
            for row in rows:
                cat_val = row[pie_col].strip()
                num_val = _to_float_safe(row.get(revenue_col, ""))
                if cat_val and num_val is not None:
                    cat_totals[cat_val] += num_val

            if cat_totals:
                sorted_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)
                labels = [c[0] for c in sorted_cats[:8]]
                sizes = [c[1] for c in sorted_cats[:8]]
                colors = ["#e94560", "#0f3460", "#533483", "#16213e", "#1a1a2e", "#e94560aa", "#0f3460aa", "#533483aa"]

                fig, ax = plt.subplots(figsize=(10, 8))
                wedges, texts, autotexts = ax.pie(
                    sizes, labels=labels, autopct="%1.1f%%",
                    colors=colors[:len(labels)],
                    startangle=90,
                    textprops={"color": "#eee", "fontsize": 11},
                    wedgeprops={"edgecolor": "#1a1a2e", "linewidth": 2},
                )
                for autotext in autotexts:
                    autotext.set_fontweight("bold")
                ax.set_title(f"{revenue_col} by {pie_col}", fontsize=16, fontweight="bold", pad=15)
                plt.tight_layout()

                chart2_path = os.path.join(output_dir, "chart_distribution.png")
                plt.savefig(chart2_path, dpi=150, bbox_inches="tight")
                plt.close()
                chart_files.append(chart2_path)
                charts_metadata.append({
                    "id": "chart_2",
                    "name": f"{pie_col} Distribution",
                    "type": "pie",
                    "title": f"{revenue_col} by {pie_col}",
                    "file": chart2_path,
                    "description": f"Shows {revenue_col} distribution across {pie_col} categories",
                })

        # ── Chart 3: Top Items Bar Chart ──
        bar_col = None
        for col in categorical_cols:
            if any(kw in col.lower() for kw in ("product", "item", "name", "title")):
                bar_col = col
                break
        if not bar_col and len(categorical_cols) > 1:
            bar_col = categorical_cols[1]
        elif not bar_col and categorical_cols:
            bar_col = categorical_cols[0]

        if bar_col and bar_col != pie_col:
            item_totals = defaultdict(float)
            for row in rows:
                item = row[bar_col].strip()
                val = _to_float_safe(row.get(revenue_col, ""))
                if item and val is not None:
                    item_totals[item] += val

            if item_totals:
                top_items = sorted(item_totals.items(), key=lambda x: x[1], reverse=True)[:10]
                names = [t[0] for t in top_items]
                values = [t[1] for t in top_items]

                fig, ax = plt.subplots(figsize=(12, 7))
                bars = ax.barh(names[::-1], values[::-1], color="#e94560", edgecolor="#0f3460", linewidth=1.5, height=0.6)
                ax.set_title(f"Top {len(names)} {bar_col} by {revenue_col}", fontsize=16, fontweight="bold", pad=15)
                ax.set_xlabel(revenue_col, fontsize=12)
                ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
                ax.grid(True, axis="x", linestyle="--", alpha=0.3)

                # Add value labels
                for bar_item, val in zip(bars, values[::-1]):
                    ax.text(val + max(values) * 0.01, bar_item.get_y() + bar_item.get_height() / 2,
                            f"${val:,.0f}", va="center", fontsize=10, color="#eee")
                plt.tight_layout()

                chart3_path = os.path.join(output_dir, "chart_top_items.png")
                plt.savefig(chart3_path, dpi=150, bbox_inches="tight")
                plt.close()
                chart_files.append(chart3_path)
                charts_metadata.append({
                    "id": "chart_3",
                    "name": f"Top {bar_col}",
                    "type": "bar",
                    "title": f"Top {len(names)} {bar_col} by {revenue_col}",
                    "file": chart3_path,
                    "description": f"Ranking of {bar_col} by total {revenue_col}",
                })

        # ── Chart 4: Secondary category bar chart ──
        if pie_col:
            cat_totals_2 = defaultdict(float)
            for row in rows:
                cat_val = row[pie_col].strip()
                val = _to_float_safe(row.get(revenue_col, ""))
                if cat_val and val is not None:
                    cat_totals_2[cat_val] += val

            if cat_totals_2:
                sorted_cats_2 = sorted(cat_totals_2.items(), key=lambda x: x[1], reverse=True)
                cat_names = [c[0] for c in sorted_cats_2]
                cat_vals = [c[1] for c in sorted_cats_2]

                fig, ax = plt.subplots(figsize=(10, 6))
                bar_colors = ["#e94560", "#0f3460", "#533483", "#16213e", "#e94560aa", "#0f3460aa"]
                ax.bar(cat_names, cat_vals, color=bar_colors[:len(cat_names)],
                       edgecolor="#1a1a2e", linewidth=1.5, width=0.6)
                ax.set_title(f"{revenue_col} by {pie_col} (Comparison)", fontsize=16, fontweight="bold", pad=15)
                ax.set_ylabel(revenue_col, fontsize=12)
                ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
                ax.grid(True, axis="y", linestyle="--", alpha=0.3)

                # Add value labels on bars
                for i, (name, val) in enumerate(zip(cat_names, cat_vals)):
                    ax.text(i, val + max(cat_vals) * 0.02, f"${val:,.0f}",
                            ha="center", fontsize=10, color="#eee", fontweight="bold")
                plt.xticks(rotation=45, ha="right")
                plt.tight_layout()

                chart4_path = os.path.join(output_dir, "chart_category_comparison.png")
                plt.savefig(chart4_path, dpi=150, bbox_inches="tight")
                plt.close()
                chart_files.append(chart4_path)
                charts_metadata.append({
                    "id": "chart_4",
                    "name": f"{pie_col} Comparison",
                    "type": "bar",
                    "title": f"{revenue_col} by {pie_col} (Comparison)",
                    "file": chart4_path,
                    "description": f"Side-by-side comparison of {revenue_col} across {pie_col}",
                })

    # ── Create Excel Workbook ──
    excel_path = None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # -- Sheet 1: Data --
        ws_data = wb.active
        ws_data.title = "Data"
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header, "")
                # Try to write as number
                num = _to_float_safe(val)
                cell = ws_data.cell(row=row_idx, column=col_idx, value=num if num is not None else val)
                cell.border = thin_border

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            max_width = max(len(str(headers[col_idx - 1])), 12)
            ws_data.column_dimensions[get_column_letter(col_idx)].width = max_width + 2

        # -- Sheet 2: Summary --
        ws_summary = wb.create_sheet("Summary")
        title_font = Font(name="Arial", size=14, bold=True, color="0F3460")
        metric_font = Font(name="Arial", size=11, bold=True)
        value_font = Font(name="Arial", size=11)

        ws_summary.cell(row=1, column=1, value="📊 Analysis Summary").font = title_font
        ws_summary.cell(row=2, column=1, value=f"Total Records: {len(rows)}")
        ws_summary.cell(row=3, column=1, value=f"Total Columns: {len(headers)}")

        # Summary metrics
        row_num = 5
        ws_summary.cell(row=row_num, column=1, value="Key Metrics").font = Font(size=12, bold=True, color="E94560")
        row_num += 1

        if revenue_col:
            revenue_values = [_to_float_safe(row.get(revenue_col, "")) for row in rows]
            revenue_values = [v for v in revenue_values if v is not None]
            if revenue_values:
                metrics = [
                    (f"Total {revenue_col}", f"${sum(revenue_values):,.2f}"),
                    (f"Average {revenue_col}", f"${statistics.mean(revenue_values):,.2f}"),
                    (f"Median {revenue_col}", f"${statistics.median(revenue_values):,.2f}"),
                    (f"Max {revenue_col}", f"${max(revenue_values):,.2f}"),
                    (f"Min {revenue_col}", f"${min(revenue_values):,.2f}"),
                ]
                for label, value in metrics:
                    ws_summary.cell(row=row_num, column=1, value=label).font = metric_font
                    ws_summary.cell(row=row_num, column=2, value=value).font = value_font
                    row_num += 1

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # -- Sheet 3: Insights --
        ws_insights = wb.create_sheet("Insights")
        ws_insights.cell(row=1, column=1, value="💡 Key Insights & Recommendations").font = title_font
        ws_insights.cell(row=3, column=1, value="Note: Run the full analysis pipeline for AI-generated insights").font = Font(size=11, italic=True, color="888888")
        ws_insights.column_dimensions["A"].width = 60

        # Save
        excel_path = os.path.join(output_dir, "analysis_report.xlsx")
        wb.save(excel_path)

    except ImportError:
        excel_path = None

    # ── Build result ──
    result = {
        "status": "success",
        "charts_created": len(chart_files),
        "charts": charts_metadata,
        "excel_report": excel_path,
        "output_directory": output_dir,
        "sheet_structure": {
            "Data": f"Complete cleaned dataset ({len(rows)} rows)",
            "Summary": "Key metrics and KPIs",
            "Insights": "Findings and recommendations",
        },
        "summary_metrics": [],
        "formatting": {
            "color_scheme": "dark_professional (navy + crimson)",
            "chart_style": "dark theme with gradient fills",
            "font": "Arial 11pt",
        },
    }

    # Add summary metrics
    if revenue_col:
        revenue_values = [_to_float_safe(row.get(revenue_col, "")) for row in rows]
        revenue_values = [v for v in revenue_values if v is not None]
        if revenue_values:
            result["summary_metrics"] = [
                {"label": f"Total {revenue_col}", "value": f"${sum(revenue_values):,.2f}", "format": "currency"},
                {"label": "Total Transactions", "value": str(len(rows)), "format": "number"},
                {"label": f"Average {revenue_col}", "value": f"${statistics.mean(revenue_values):,.2f}", "format": "currency"},
                {"label": f"Max {revenue_col}", "value": f"${max(revenue_values):,.2f}", "format": "currency"},
            ]

    return json.dumps(result, indent=2)
